from tkinter import *
from openpyxl import Workbook, load_workbook
import re
import pyperclip
from bs4 import BeautifulSoup as Soup
import requests
import configparser
import textwrap


config = configparser.ConfigParser()
secrets = r"c:/source/secrets"
config.read(secrets)
apikey = config.get("configuration", "PRODUCTION")


class AutocompleteEntry(Entry):
    def __init__(self, lista, *args, **kwargs):
        Entry.__init__(self, *args, **kwargs)
        self.lista = lista   
        self.var = self["textvariable"]
        if self.var == '':
            self.var = self["textvariable"] = StringVar()    
        
        self.mms_label  = StringVar()

        self.title_body = Label(root, width=width, height=6, text="", textvariable=self.mms_label)
        self.title_body.grid(row=1, columnspan=3)
        self.var.trace('w', self.changed)
        self.bind("<Right>", self.selection)
        self.bind("<Up>", self.up)
        self.bind("<Down>", self.down)
        
        self.lb_up = False

    def changed(self, name, index, mode):  

        if self.var.get() == '':
            self.lb.destroy()
            self.lb_up = False
        else:
            words = self.comparison()
            if words:            
                if not self.lb_up:
                    self.lb = Listbox()
                    self.lb.bind("<Double-Button-1>", self.selection)
                    self.lb.bind("<Right>", self.selection)
                    self.lb.place(x=self.winfo_x(), y=self.winfo_y()+self.winfo_height())
                    self.lb_up = True
                
                self.lb.delete(0, END)
                for w in words:
                    self.lb.insert(END,w)
            else:
                if self.lb_up:
                    self.lb.destroy()
                    self.lb_up = False
        
    def selection(self, event):

        if self.lb_up:
            self.var.set(self.lb.get(ACTIVE))
            self.lb.destroy()
            self.lb_up = False
            self.icursor(END)
            self.mms_label.set( f"{titles_lookup[self.var.get()]['mms']}") 

    def up(self, event):
        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != '0':                
                self.lb.selection_clear(first=index)
                index = str(int(index)-1)                
                self.lb.selection_set(first=index)
                self.lb.activate(index) 

    def down(self, event):
        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != END:                        
                self.lb.selection_clear(first=index)
                index = str(int(index)+1)        
                self.lb.selection_set(first=index)
                self.lb.activate(index) 
                print ("1", self.var.get())

    def comparison(self):
        pattern = re.compile('.*' + self.var.get() + '.*', re.IGNORECASE)
        return [w for w in self.lista if re.match(pattern, w)]


def get_title_and_mms_lookup():
	titles_lookup = {}
	mms_lookup = {}
	len_data = 11
	my_data = []
	try:
		wb = load_workbook(r"G:\Fileplan\Bib_Services\Non-Clio_formats\Acquisitions Team\bulk item ingest\titles_reference.xlsx", data_only=True)
	except:	
		wb = load_workbook("titles_reference.xlsx", data_only=True)
	ws = wb.active
	row_counter = 0
	cell_counter = 0
	for i, row in enumerate(ws):
		my_row = []
		row_counter  += 1
		if row_counter == 1:
			pass
		else:
			for cell in row[0:len_data]:
				try:
					my_row.append(int(cell.value))
				except:
					my_row.append(cell.value)

			if my_row != [None for x in range(len_data)]:
				my_data.append(my_row[0:len_data])

	for row in my_data:
		if row != [None, None, None, None, None, None, None]:
			### these are all grubby excel fixes. :( 
			mms = row[1]
			if str(mms).endswith("0"):
				mms+=6
			atl = row[4]
			if str(atl).endswith("00"):
				atl += 36
			wn  = row[3]
			if str(wn).endswith("00"):
				wn += 36
			
			titles_lookup[row[0]] = {"mms":mms, "ATL":atl, "WN":wn, "pol":row[2], "title":row[0], 'signed_off':row[6]}
			mms_lookup[mms] = {"mms":mms, "ATL":atl, "WN":wn, "pol":row[2], "title":row[0], 'signed_off':row[6]}

	return titles_lookup, mms_lookup

def get_holding_note(mms):
	my_pol = mms_lookup[mms]['pol']
	url = f"https://api-ap.hosted.exlibrisgroup.com/almaws/v1/acq/po-lines/{my_pol}?apikey={apikey}"
	r = requests.get(url).text
	soup = Soup(r, 'lxml')
	my_note = soup.find("receiving_note").text
	
	if my_note != None: 
		my_note = my_note.strip()

	if my_note == "" or my_note == None:
		my_note = "No Note"

	my_note = my_note.replace(". ", ".\n")
	my_note_lines = my_note.split("\n")
	note_cleaned_lines = [] 
	for line in my_note_lines:
		my_line = textwrap.wrap(line, width, break_long_words=False)
		note_cleaned_lines.append("\n".join(my_line))

	my_note = "\n".join(note_cleaned_lines)
	return (my_note)

def get_and_update_with_rcv_note():
	try:
		mms = titles_lookup[entry.var.get()]['mms']
		my_note = get_holding_note(mms)
		entry.mms_label.set( f"{my_note}") 
	except KeyError:
		entry.mms_label.set( f"\n\tNo MMS selected")


def clipboard():
	try:
		pyperclip.copy(titles_lookup[entry.var.get()]['mms'])
	except KeyError:
		entry.mms_label.set( f"\n\tNo MMS selected")

def quit():
	root.destroy()



if __name__ == '__main__':
	width = 60
	root = Tk()
	titles_lookup, mms_lookup = get_title_and_mms_lookup()
	lista = list(titles_lookup.keys())
	entry = AutocompleteEntry(lista, root)
	entry.grid(row=0, column=0, columnspan=3)
	Button(root, text='Copy MMS to clipboard', command=clipboard).grid(row=2, column=0)
	Button(root, text='Get Rcv Note', command=get_and_update_with_rcv_note).grid(row=2, column=1)
	Button(root, text='Quit', command=quit).grid(row=2, column=2)
	root.lift()
	root.call('wm', 'attributes', '.', '-topmost', True)
	root.mainloop()
